"""
Service for factura processing.
Handles both legacy Excel workflow and new pending invoices workflow.
"""
import io
import re
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.infrastructure.odoo import OdooClient
from app.models import (
    PendingInvoice,
    PendingInvoiceItem,
    InvoiceHistory,
    InvoiceHistoryItem,
    InvoiceStatus
)
from app.schemas.auth import UserInfo
from app.schemas.invoice import (
    InvoiceUploadResponse,
    PendingInvoiceListResponse,
    PendingInvoiceResponse,
    InvoiceItemResponse,
    InvoiceSyncResponse,
    InvoiceHistoryListResponse,
    InvoiceHistoryResponse,
    InvoiceHistoryItemResponse,
    ProductPreview,
    InvoicePreview,
    InvoicePreviewResponse
)
from app.core.constants import UserRole, OdooModel
from app.utils.timezone import get_ecuador_now
from .utils import extract_productos_from_xml, extract_productos_preview_from_xml, create_unified_xml, update_xml_with_barcodes, update_xml_with_barcodes_consolidated


logger = logging.getLogger(__name__)


class FacturaService:
    """Service for processing facturas and generating Excel."""

    def __init__(self, db: Session = None, odoo_client: OdooClient = None):
        """
        Initialize factura service.

        Args:
            db: Database session for persistence
            odoo_client: Odoo client for syncing
        """
        self.db = db
        self.odoo_client = odoo_client

    # ========================================================================
    # NEW METHODS - PENDING INVOICES WORKFLOW
    # ========================================================================

    def upload_and_create_pending_invoices(
        self,
        xml_files: List[Dict[str, str]],
        uploaded_by: UserInfo,
        barcode_source: str = 'codigoAuxiliar'
    ) -> InvoiceUploadResponse:
        """
        Admin uploads XML invoices from SRI and creates pending invoices.

        Args:
            xml_files: List of dicts with 'filename' and 'content'
            uploaded_by: User uploading the invoices
            barcode_source: Which XML field to use as barcode ('codigoPrincipal' or 'codigoAuxiliar')

        Returns:
            InvoiceUploadResponse with summary
        """
        if not self.db:
            raise ValueError("Database session required")

        invoice_ids = []
        total_products = 0

        for xml_data in xml_files:
            try:
                filename = xml_data['filename']
                content = xml_data['content']

                # Extract products from XML with barcode source preference
                productos = extract_productos_from_xml(content, barcode_source=barcode_source)

                # Extract invoice metadata
                metadata = self._extract_invoice_metadata(content)

                # Create pending invoice
                pending_invoice = PendingInvoice(
                    invoice_number=metadata.get('invoice_number'),
                    supplier_name=metadata.get('supplier_name'),
                    invoice_date=metadata.get('invoice_date'),
                    uploaded_by_id=uploaded_by.user_id if uploaded_by.user_id else None,
                    uploaded_by_username=uploaded_by.username,
                    xml_filename=filename,
                    xml_content=content,
                    barcode_source=barcode_source,
                    status=InvoiceStatus.PENDIENTE_REVISION
                )

                self.db.add(pending_invoice)
                self.db.flush()  # Get ID

                # Create items
                for producto in productos:
                    item = PendingInvoiceItem(
                        invoice_id=pending_invoice.id,
                        codigo_original=producto['codigo'],
                        product_name=producto['descripcion'],
                        quantity=producto.get('cantidad', 0),
                        cantidad_original=producto.get('cantidad', 0),
                        barcode=None,  # Bodeguero will fill this
                        unit_price=producto.get('precio_unitario'),
                        total_price=producto.get('precio_total'),
                        modified_by_bodeguero=False
                    )
                    self.db.add(item)

                self.db.commit()

                invoice_ids.append(pending_invoice.id)
                total_products += len(productos)

                logger.info(f"Created pending invoice {pending_invoice.id} from {filename} with {len(productos)} products")

            except Exception as e:
                self.db.rollback()
                logger.error(f"Error processing {xml_data.get('filename')}: {str(e)}")
                raise

        return InvoiceUploadResponse(
            success=True,
            message=f"Successfully created {len(invoice_ids)} pending invoices",
            invoices_created=len(invoice_ids),
            invoice_ids=invoice_ids,
            total_products=total_products
        )

    def preview_invoices(
        self,
        xml_files: List[Dict[str, str]]
    ) -> InvoicePreviewResponse:
        """
        Preview XML invoices without creating database records.

        Extracts all products with BOTH barcode fields visible for user decision.
        Does NOT write to database - this is a read-only preview operation.

        Args:
            xml_files: List of dicts with 'filename' and 'content'

        Returns:
            InvoicePreviewResponse with all products from all files
        """
        previews = []
        total_products = 0

        for xml_data in xml_files:
            try:
                filename = xml_data['filename']
                content = xml_data['content']

                # Extract products with BOTH barcode fields
                productos = extract_productos_preview_from_xml(content)

                # Extract invoice metadata
                metadata = self._extract_invoice_metadata(content)

                # Build preview object
                preview = InvoicePreview(
                    filename=filename,
                    invoice_number=metadata.get('invoice_number'),
                    supplier_name=metadata.get('supplier_name'),
                    invoice_date=metadata.get('invoice_date'),
                    products=[
                        ProductPreview(
                            codigo_principal=p['codigo_principal'],
                            codigo_auxiliar=p['codigo_auxiliar'],
                            descripcion=p['descripcion'],
                            cantidad=p['cantidad'],
                            precio_unitario=p.get('precio_unitario'),
                            precio_total=p.get('precio_total')
                        )
                        for p in productos
                    ],
                    total_products=len(productos)
                )

                previews.append(preview)
                total_products += len(productos)

                logger.info(f"Preview generated for {filename} with {len(productos)} products")

            except Exception as e:
                logger.error(f"Error previewing {xml_data.get('filename', 'unknown')}: {str(e)}")
                # Continue with other files
                continue

        return InvoicePreviewResponse(
            success=True,
            message=f"Preview generated for {len(previews)} file(s)",
            previews=previews,
            total_files=len(previews),
            total_products=total_products
        )

    def get_pending_invoices(
        self,
        user: UserInfo,
        status: Optional[str] = None
    ) -> PendingInvoiceListResponse:
        """
        Get pending invoices filtered by role.

        Args:
            user: Current user
            status: Optional status filter

        Returns:
            List of pending invoices
        """
        if not self.db:
            raise ValueError("Database session required")

        query = self.db.query(PendingInvoice)

        # Filter by role
        if user.role == UserRole.BODEGUERO:
            # Bodeguero only sees PENDIENTE_REVISION and EN_REVISION
            query = query.filter(
                PendingInvoice.status.in_([
                    InvoiceStatus.PENDIENTE_REVISION,
                    InvoiceStatus.EN_REVISION
                ])
            )

        # Additional status filter
        if status:
            query = query.filter(PendingInvoice.status == status)

        # Order by created_at desc
        query = query.order_by(PendingInvoice.created_at.desc())

        invoices = query.all()

        # Convert to response with price filtering
        invoice_responses = [
            self._invoice_to_response(invoice, user)
            for invoice in invoices
        ]

        return PendingInvoiceListResponse(
            invoices=invoice_responses,
            total=len(invoice_responses)
        )

    def get_pending_invoice_by_id(
        self,
        invoice_id: int,
        user: UserInfo
    ) -> PendingInvoiceResponse:
        """
        Get single pending invoice with items.

        Args:
            invoice_id: Invoice ID
            user: Current user

        Returns:
            Invoice detail with price filtering
        """
        if not self.db:
            raise ValueError("Database session required")

        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        return self._invoice_to_response(invoice, user)

    def update_invoice_item(
        self,
        invoice_id: int,
        item_id: int,
        quantity: float,
        barcode: Optional[str],
        user: UserInfo
    ) -> InvoiceItemResponse:
        """
        Bodeguero updates quantity and/or barcode for an item.

        Args:
            invoice_id: Invoice ID
            item_id: Item ID
            quantity: New quantity
            barcode: New barcode (optional)
            user: Current user (must be bodeguero)

        Returns:
            Updated item
        """
        if not self.db:
            raise ValueError("Database session required")

        # Verify user is bodeguero
        if user.role != UserRole.BODEGUERO:
            raise ValueError("Only bodegueros can update invoice items")

        # Get invoice
        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Verify status
        if invoice.status not in [InvoiceStatus.PENDIENTE_REVISION, InvoiceStatus.EN_REVISION]:
            raise ValueError(f"Cannot update invoice in status {invoice.status}")

        # Get item
        item = self.db.query(PendingInvoiceItem).filter(
            PendingInvoiceItem.id == item_id,
            PendingInvoiceItem.invoice_id == invoice_id
        ).first()

        if not item:
            raise ValueError(f"Item {item_id} not found in invoice {invoice_id}")

        # Update item
        item.quantity = quantity
        if barcode is not None:
            item.barcode = barcode
        item.modified_by_bodeguero = True

        # Update invoice status to EN_REVISION if not already
        if invoice.status == InvoiceStatus.PENDIENTE_REVISION:
            invoice.status = InvoiceStatus.EN_REVISION

        self.db.commit()

        logger.info(f"Updated item {item_id} in invoice {invoice_id} by {user.username}")

        return self._item_to_response(item, user)

    def update_item_sale_price(
        self,
        invoice_id: int,
        item_id: int,
        manual_sale_price: Optional[float],
        user: UserInfo
    ) -> InvoiceItemResponse:
        """
        Admin updates manual sale price for an item.

        Args:
            invoice_id: Invoice ID
            item_id: Item ID
            manual_sale_price: Manual sale price with IVA (null to revert to calculated)
            user: Current user (must be admin)

        Returns:
            Updated item
        """
        if not self.db:
            raise ValueError("Database session required")

        # Verify user is admin
        if user.role != UserRole.ADMIN:
            raise ValueError("Only admins can update sale prices")

        # Get invoice
        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Verify status (can only edit prices in CORREGIDA or PARCIALMENTE_SINCRONIZADA)
        if invoice.status not in [InvoiceStatus.CORREGIDA, InvoiceStatus.PARCIALMENTE_SINCRONIZADA]:
            raise ValueError(f"Cannot update prices for invoice in status {invoice.status}")

        # Get item
        item = self.db.query(PendingInvoiceItem).filter(
            PendingInvoiceItem.id == item_id,
            PendingInvoiceItem.invoice_id == invoice_id
        ).first()

        if not item:
            raise ValueError(f"Item {item_id} not found in invoice {invoice_id}")

        # Update manual sale price
        item.manual_sale_price = manual_sale_price

        self.db.commit()

        logger.info(f"Updated manual sale price for item {item_id} in invoice {invoice_id} by {user.username}: {manual_sale_price}")

        return self._item_to_response(item, user)

    def admin_update_invoice_item(
        self,
        invoice_id: int,
        item_id: int,
        quantity: Optional[float],
        barcode: Optional[str],
        product_name: Optional[str],
        user: UserInfo
    ) -> InvoiceItemResponse:
        """
        Admin updates invoice item (can edit quantity, barcode, and product name).

        Args:
            invoice_id: Invoice ID
            item_id: Item ID
            quantity: New quantity (optional)
            barcode: New barcode (optional)
            product_name: New product name (optional)
            user: Current user (must be admin)

        Returns:
            Updated item
        """
        if not self.db:
            raise ValueError("Database session required")

        # Verify user is admin
        if user.role != UserRole.ADMIN:
            raise ValueError("Only admins can use this endpoint")

        # Get invoice
        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Verify status
        if invoice.status not in [InvoiceStatus.CORREGIDA, InvoiceStatus.PARCIALMENTE_SINCRONIZADA]:
            raise ValueError(f"Cannot update items for invoice in status {invoice.status}")

        # Get item
        item = self.db.query(PendingInvoiceItem).filter(
            PendingInvoiceItem.id == item_id,
            PendingInvoiceItem.invoice_id == invoice_id
        ).first()

        if not item:
            raise ValueError(f"Item {item_id} not found in invoice {invoice_id}")

        # Update fields if provided
        if quantity is not None:
            item.quantity = quantity

        if barcode is not None:
            item.barcode = barcode if barcode.strip() else None

        if product_name is not None:
            item.product_name = product_name

        self.db.commit()

        logger.info(f"Admin {user.username} updated item {item_id} in invoice {invoice_id}: qty={quantity}, barcode={barcode}, name={product_name}")

        return self._item_to_response(item, user)

    def exclude_invoice_item(
        self,
        invoice_id: int,
        item_id: int,
        is_excluded: bool,
        reason: Optional[str],
        user: UserInfo
    ) -> InvoiceItemResponse:
        """
        Exclude or include an item from sync (admin only).

        Args:
            invoice_id: Invoice ID
            item_id: Item ID
            is_excluded: True to exclude, False to include
            reason: Optional reason for exclusion
            user: Current user (must be admin)

        Returns:
            Updated item
        """
        if not self.db:
            raise ValueError("Database session required")

        # Verify user is admin
        if user.role != UserRole.ADMIN:
            raise ValueError("Only admins can exclude items")

        # Get invoice
        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Verify status
        if invoice.status not in [InvoiceStatus.CORREGIDA, InvoiceStatus.PARCIALMENTE_SINCRONIZADA]:
            raise ValueError(f"Cannot exclude items for invoice in status {invoice.status}")

        # Get item
        item = self.db.query(PendingInvoiceItem).filter(
            PendingInvoiceItem.id == item_id,
            PendingInvoiceItem.invoice_id == invoice_id
        ).first()

        if not item:
            raise ValueError(f"Item {item_id} not found in invoice {invoice_id}")

        # Update exclusion
        item.is_excluded = is_excluded
        item.excluded_reason = reason if is_excluded else None

        self.db.commit()

        logger.info(f"Admin {user.username} {'excluded' if is_excluded else 'included'} item {item_id} in invoice {invoice_id}: {reason}")

        return self._item_to_response(item, user)

    def submit_invoice(
        self,
        invoice_id: int,
        user: UserInfo,
        notes: Optional[str]
    ) -> PendingInvoiceResponse:
        """
        Bodeguero marks invoice as completed.

        Args:
            invoice_id: Invoice ID
            user: Current user (must be bodeguero)
            notes: Optional notes

        Returns:
            Updated invoice
        """
        if not self.db:
            raise ValueError("Database session required")

        # Verify user is bodeguero
        if user.role != UserRole.BODEGUERO:
            raise ValueError("Only bodegueros can submit invoices")

        # Get invoice
        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Update invoice
        invoice.status = InvoiceStatus.CORREGIDA
        invoice.submitted_at = get_ecuador_now().replace(tzinfo=None)
        invoice.submitted_by = user.username
        if notes:
            invoice.notes = notes

        self.db.commit()

        logger.info(f"Invoice {invoice_id} submitted by {user.username}")

        return self._invoice_to_response(invoice, user)

    def _transform_items_to_product_format(
        self,
        items: List[PendingInvoiceItem],
        profit_margin: float,
        apply_iva: bool,
        quantity_mode: str
    ) -> List[Dict[str, Any]]:
        """
        Transform PendingInvoiceItems to ProductService format.
        Applies consolidation and price calculation logic from XML parser.

        Args:
            items: Items from pending invoice
            profit_margin: Profit margin (0.5 = 50%)
            apply_iva: Whether to calculate price with IVA
            quantity_mode: 'add' or 'replace' stock mode

        Returns:
            List of products in format expected by ProductService.sync_products_bulk()
        """
        from app.utils.formatters import (
            apply_profit_margin,
            calculate_price_without_iva,
            round_to_half_dollar
        )

        # 0. Filter out excluded items (they should not be synced)
        items_to_process = [item for item in items if not item.is_excluded]

        # 1. Consolidate items by codigo_original
        # Use barcode if assigned, otherwise fall back to codigo_original
        # (the main code often IS the barcode, workers just don't fill the barcode field)
        consolidated = {}

        for item in items_to_process:
            # Use barcode if available, otherwise use codigo_original as barcode
            effective_barcode = item.barcode or item.codigo_original

            if effective_barcode not in consolidated:
                consolidated[effective_barcode] = {
                    'items': [],
                    'total_quantity': 0.0,
                    'total_amount': 0.0,
                    'name': item.product_name,
                    'manual_sale_prices': []  # Track manual prices for this barcode
                }

            # Use total_price if exists, otherwise calculate
            line_total = item.total_price or (item.quantity * item.unit_price)

            consolidated[effective_barcode]['items'].append(item)
            consolidated[effective_barcode]['total_quantity'] += item.quantity
            consolidated[effective_barcode]['total_amount'] += line_total

            # Track manual sale price if set
            if item.manual_sale_price is not None:
                consolidated[effective_barcode]['manual_sale_prices'].append(item.manual_sale_price)

        # 2. Calculate real cost and transform to product format
        mapped_products = []

        for barcode, data in consolidated.items():
            # Calculate weighted average unit cost
            if data['total_quantity'] > 0:
                real_unit_cost = data['total_amount'] / data['total_quantity']
            else:
                real_unit_cost = 0.0

            # Handle 100% discount products (cost = 0)
            if real_unit_cost == 0.0 and data['items']:
                avg_unit_price = sum(i.unit_price for i in data['items']) / len(data['items'])
                real_unit_cost = avg_unit_price

            # 3. Calculate sale price
            # Check if admin set a manual price
            manual_prices = data.get('manual_sale_prices', [])
            if manual_prices and all(p == manual_prices[0] for p in manual_prices):
                # All items have same manual price - use it
                display_price = manual_prices[0]

                # Calculate price without IVA for Odoo
                if apply_iva:
                    sale_price = calculate_price_without_iva(display_price)
                else:
                    sale_price = display_price
                    display_price = None
            else:
                # No manual price - calculate automatically (like xml_parser.py:574-632)
                # Apply profit margin
                price_with_margin = apply_profit_margin(real_unit_cost, profit_margin)

                # Round to next $0.50
                display_price = round_to_half_dollar(price_with_margin)

                # Calculate price without IVA for Odoo
                if apply_iva:
                    sale_price = calculate_price_without_iva(display_price)
                else:
                    sale_price = display_price
                    display_price = None

            # 4. Create product in expected format
            mapped_product = {
                "name": data['name'],
                "qty_available": data['total_quantity'],
                "barcode": barcode,
                "standard_price": real_unit_cost,      # Cost
                "list_price": sale_price,              # Sale price without IVA
                "display_price": display_price,        # Sale price with IVA
                "type": "storable",
                "tracking": "none",
                "available_in_pos": True,
                "quantity_mode": quantity_mode
            }

            mapped_products.append(mapped_product)

        return mapped_products

    def sync_invoice_to_odoo(
        self,
        invoice_id: int,
        user: UserInfo,
        notes: Optional[str],
        item_ids: Optional[List[int]] = None
    ) -> InvoiceSyncResponse:
        """
        Admin synchronizes invoice to Odoo.
        Supports partial sync by selecting specific items.

        Args:
            invoice_id: Invoice ID
            user: Current user (must be admin)
            notes: Optional admin notes
            item_ids: Optional list of item IDs to sync. If None, syncs all items.

        Returns:
            Sync result with errors
        """
        if not self.db:
            raise ValueError("Database session required")

        if not self.odoo_client:
            raise ValueError("Odoo client required")

        # Verify user is admin
        if user.role != UserRole.ADMIN:
            raise ValueError("Only admins can sync invoices")

        # Get invoice
        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Verify status (can sync if CORREGIDA or PARCIALMENTE_SINCRONIZADA)
        if invoice.status not in [InvoiceStatus.CORREGIDA, InvoiceStatus.PARCIALMENTE_SINCRONIZADA]:
            raise ValueError(
                f"Invoice must be in CORREGIDA or PARCIALMENTE_SINCRONIZADA status to sync (current: {invoice.status})"
            )

        try:
            # 1. Filter items if item_ids provided (partial sync)
            items_to_sync = invoice.items
            if item_ids is not None:
                items_to_sync = [item for item in invoice.items if item.id in item_ids]
                if not items_to_sync:
                    raise ValueError(f"No valid items found with IDs: {item_ids}")
                logger.info(f"Partial sync: selected {len(items_to_sync)} of {len(invoice.items)} items")
            else:
                logger.info(f"Full sync: processing all {len(invoice.items)} items")

            # 2. Transform items to product format
            mapped_products = self._transform_items_to_product_format(
                items=items_to_sync,
                profit_margin=invoice.profit_margin,
                apply_iva=invoice.apply_iva,
                quantity_mode=invoice.quantity_mode
            )

            logger.info(f"Consolidated to {len(mapped_products)} unique products")

            # 2. Use ProductService to sync (reuses ALL existing logic)
            from app.features.products.service import ProductService

            product_service = ProductService(
                odoo_client=self.odoo_client,
                db=self.db
            )

            # 3. Sync all products (creates/updates products, prices, and stock)
            sync_response = product_service.sync_products_bulk(
                products=mapped_products,
                username=user.username,
                xml_filename=invoice.xml_filename or f"invoice_{invoice.invoice_number}.xml",
                xml_provider=invoice.supplier_name,
                profit_margin=invoice.profit_margin,
                quantity_mode=invoice.quantity_mode,
                apply_iva=invoice.apply_iva,
                xml_content=invoice.xml_content
            )

            # 4. Update items with sync results
            # Map barcode -> items for updating
            items_by_barcode = {}
            for item in invoice.items:
                if item.barcode:
                    if item.barcode not in items_by_barcode:
                        items_by_barcode[item.barcode] = []
                    items_by_barcode[item.barcode].append(item)

            # Update items based on sync results
            errors = []
            successful_items = []
            failed_items = []

            for result in sync_response.results:
                if result.barcode in items_by_barcode:
                    for item in items_by_barcode[result.barcode]:
                        if result.success:
                            item.product_id = result.product_id
                            item.sync_success = True
                            item.sync_error = None
                            successful_items.append(item)
                        else:
                            item.sync_success = False
                            item.sync_error = result.message
                            failed_items.append({'item': item, 'error': result.message})
                            errors.append(f"{item.product_name} ({result.barcode}): {result.message}")

        except Exception as e:
            logger.error(f"Error syncing invoice {invoice_id}: {str(e)}")
            self.db.rollback()
            raise

        # Update invoice status based on sync completeness
        # Count how many items are successfully synced
        total_synced = sum(1 for item in invoice.items if item.sync_success)
        total_items = len(invoice.items)

        if total_synced == total_items:
            # All items synced
            invoice.status = InvoiceStatus.SINCRONIZADA
            logger.info(f"Invoice {invoice_id} fully synced: {total_synced}/{total_items} items")
        elif total_synced > 0:
            # Some items synced
            invoice.status = InvoiceStatus.PARCIALMENTE_SINCRONIZADA
            logger.info(f"Invoice {invoice_id} partially synced: {total_synced}/{total_items} items")
        else:
            # No items synced (keep current status)
            logger.warning(f"Invoice {invoice_id}: no items synced successfully")

        invoice.synced_at = get_ecuador_now().replace(tzinfo=None)
        invoice.synced_by = user.username
        if notes:
            invoice.notes = (invoice.notes or "") + f"\nAdmin: {notes}"

        # Create history record
        self._create_history_record(
            invoice,
            successful_items,
            failed_items,
            user.username
        )

        self.db.commit()

        return InvoiceSyncResponse(
            success=len(failed_items) == 0,
            message=f"Synced {len(successful_items)}/{len(invoice.items)} items successfully",
            successful_items=len(successful_items),
            failed_items=len(failed_items),
            errors=errors,
            profit_margin=invoice.profit_margin,
            quantity_mode=invoice.quantity_mode
        )

    def update_invoice_config(
        self,
        invoice_id: int,
        profit_margin: Optional[float],
        apply_iva: Optional[bool],
        quantity_mode: Optional[str],
        user: UserInfo
    ) -> PendingInvoiceResponse:
        """
        Update invoice sync configuration.

        Args:
            invoice_id: Invoice ID
            profit_margin: Profit margin (0-2)
            apply_iva: Apply IVA to sale price
            quantity_mode: 'add' or 'replace'
            user: Current user

        Returns:
            Updated invoice response
        """
        if not self.db:
            raise ValueError("Database session required")

        invoice = self.db.query(PendingInvoice).filter(
            PendingInvoice.id == invoice_id
        ).first()

        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")

        # Only allow config changes for pending/in-review/corrected invoices
        if invoice.status not in [
            InvoiceStatus.PENDIENTE_REVISION,
            InvoiceStatus.EN_REVISION,
            InvoiceStatus.CORREGIDA
        ]:
            raise ValueError(f"Cannot configure invoice in status {invoice.status}")

        # Update fields
        if profit_margin is not None:
            invoice.profit_margin = profit_margin

        if apply_iva is not None:
            invoice.apply_iva = apply_iva

        if quantity_mode is not None:
            invoice.quantity_mode = quantity_mode

        self.db.commit()
        self.db.refresh(invoice)

        logger.info(f"Invoice {invoice_id} config updated: margin={invoice.profit_margin}, iva={invoice.apply_iva}, mode={invoice.quantity_mode}")

        return self._invoice_to_response(invoice, user)

    def get_invoice_history(
        self,
        skip: int = 0,
        limit: int = 50
    ) -> InvoiceHistoryListResponse:
        """
        Get invoice history (admin only).

        Args:
            skip: Pagination offset
            limit: Page size

        Returns:
            List of history records
        """
        if not self.db:
            raise ValueError("Database session required")

        query = self.db.query(InvoiceHistory).order_by(
            InvoiceHistory.synced_at.desc()
        )

        total = query.count()
        history_records = query.offset(skip).limit(limit).all()

        history_responses = [
            self._history_to_response(record)
            for record in history_records
        ]

        return InvoiceHistoryListResponse(
            history=history_responses,
            total=total
        )

    # ========================================================================
    # HELPER METHODS
    # ========================================================================

    def _extract_invoice_metadata(self, xml_content: str) -> Dict:
        """Extract invoice number, supplier, date from XML."""
        metadata = {
            'invoice_number': None,
            'supplier_name': None,
            'invoice_date': None
        }

        try:
            # Extract invoice number (numeroAutorizacion or estab-ptoEmi-secuencial)
            auth_match = re.search(r'<numeroAutorizacion>(.*?)</numeroAutorizacion>', xml_content)
            if auth_match:
                metadata['invoice_number'] = auth_match.group(1)[:50]  # Limit to 50 chars

            # Extract supplier name (razonSocial)
            supplier_match = re.search(r'<razonSocial>(.*?)</razonSocial>', xml_content)
            if supplier_match:
                metadata['supplier_name'] = supplier_match.group(1)[:255]

            # Extract date (fechaEmision)
            date_match = re.search(r'<fechaEmision>(.*?)</fechaEmision>', xml_content)
            if date_match:
                date_str = date_match.group(1)
                # Try to parse date (format: DD/MM/YYYY)
                try:
                    metadata['invoice_date'] = datetime.strptime(date_str, '%d/%m/%Y')
                except ValueError:
                    pass

        except Exception as e:
            logger.warning(f"Error extracting invoice metadata: {str(e)}")

        return metadata

    def _create_history_record(
        self,
        pending_invoice: PendingInvoice,
        successful_items: List[PendingInvoiceItem],
        failed_items: List[Dict],
        synced_by: str
    ) -> InvoiceHistory:
        """Create history record after sync."""
        total_quantity = sum(item.quantity for item in pending_invoice.items)
        total_value = sum(
            item.total_price for item in pending_invoice.items
            if item.total_price is not None
        )

        history = InvoiceHistory(
            pending_invoice_id=pending_invoice.id,
            invoice_number=pending_invoice.invoice_number or "N/A",
            supplier_name=pending_invoice.supplier_name,
            invoice_date=pending_invoice.invoice_date,
            uploaded_by=pending_invoice.uploaded_by_username,
            synced_by=synced_by,
            synced_at=get_ecuador_now().replace(tzinfo=None),
            total_items=len(pending_invoice.items),
            successful_items=len(successful_items),
            failed_items=len(failed_items),
            total_quantity=total_quantity,
            total_value=total_value if total_value > 0 else None,
            xml_content=pending_invoice.xml_content,
            has_errors=len(failed_items) > 0,
            error_summary="; ".join([f"{f['item'].product_name}: {f['error']}" for f in failed_items]) if failed_items else None
        )

        self.db.add(history)
        self.db.flush()  # Flush to get history.id before creating items

        # Create history items
        for item in pending_invoice.items:
            # Calculate the correct unit price from total/quantity (same logic as xml_parser)
            correct_unit_price = (
                item.total_price / item.quantity
                if item.total_price and item.quantity > 0
                else item.unit_price
            )

            # Calculate the sale price that was synced to Odoo
            if item.manual_sale_price is not None:
                # Manual price was set
                sale_price_for_history = item.manual_sale_price
            else:
                # Calculate automatic price (same logic as transform method)
                from app.utils.formatters import apply_profit_margin, round_to_half_dollar
                price_with_margin = apply_profit_margin(correct_unit_price, pending_invoice.profit_margin)
                sale_price_for_history = round_to_half_dollar(price_with_margin)

            history_item = InvoiceHistoryItem(
                history_id=history.id,
                codigo_original=item.codigo_original,
                barcode=item.barcode,
                product_id=item.product_id,
                product_name=item.product_name,
                quantity=item.quantity,
                unit_price=correct_unit_price,
                sale_price=sale_price_for_history,
                total_value=item.total_price,
                success=item.sync_success or False,
                error_message=item.sync_error,
                was_modified=item.modified_by_bodeguero
            )
            self.db.add(history_item)

        return history

    def _invoice_to_response(
        self,
        invoice: PendingInvoice,
        user: UserInfo
    ) -> PendingInvoiceResponse:
        """Convert invoice to response with price filtering and consolidation."""
        # Consolidate items by codigo_original for all user roles
        items = self._consolidate_items(invoice.items, user)

        return PendingInvoiceResponse(
            id=invoice.id,
            invoice_number=invoice.invoice_number,
            supplier_name=invoice.supplier_name,
            invoice_date=invoice.invoice_date,
            status=invoice.status.value if isinstance(invoice.status, InvoiceStatus) else invoice.status,
            uploaded_by_username=invoice.uploaded_by_username,
            xml_filename=invoice.xml_filename,
            barcode_source=invoice.barcode_source,
            created_at=invoice.created_at,
            updated_at=invoice.updated_at,
            submitted_at=invoice.submitted_at,
            submitted_by=invoice.submitted_by,
            notes=invoice.notes,
            profit_margin=invoice.profit_margin,
            apply_iva=invoice.apply_iva,
            quantity_mode=invoice.quantity_mode,
            items=items,
            total_items=len(items),
            total_quantity=sum(item.quantity for item in items)
        )

    def _consolidate_items(
        self,
        items: List[PendingInvoiceItem],
        user: UserInfo
    ) -> List[InvoiceItemResponse]:
        """
        Consolidate items by codigo_original for all user views.
        Items with the same codigo_original are merged: quantities and totals are summed,
        unit_price is recalculated from total/quantity.
        Excluded items are NOT consolidated - they appear separately.
        """
        # Separate excluded items (they should appear individually)
        excluded_items = [item for item in items if item.is_excluded]
        active_items = [item for item in items if not item.is_excluded]

        # Group active items by codigo_original (main product code)
        # This ensures items with the same product code are unified even if
        # one has a barcode assigned and another doesn't yet
        consolidated = {}

        for item in active_items:
            key = item.codigo_original

            if key not in consolidated:
                consolidated[key] = {
                    'items': [],
                    'total_quantity': 0.0,
                    'total_cantidad_original': 0.0,
                    'total_price': 0.0,
                    'source_item_ids': [],
                    'first_item': item,
                    'barcode': item.barcode,
                    'any_modified': False,
                    'all_synced': True,
                    'any_failed': False,
                    'sync_errors': [],
                    'manual_sale_prices': [],
                    'product_ids': set()
                }

            data = consolidated[key]
            data['items'].append(item)
            data['total_quantity'] += item.quantity
            data['total_cantidad_original'] += item.cantidad_original
            data['total_price'] += item.total_price if item.total_price else 0.0
            data['source_item_ids'].append(item.id)

            if item.modified_by_bodeguero:
                data['any_modified'] = True

            # Track sync status
            if item.sync_success is None:
                data['all_synced'] = False
            elif item.sync_success is False:
                data['any_failed'] = True
                data['all_synced'] = False
                if item.sync_error:
                    data['sync_errors'].append(item.sync_error)

            if item.product_id:
                data['product_ids'].add(item.product_id)

            if item.manual_sale_price is not None:
                data['manual_sale_prices'].append(item.manual_sale_price)

            # Track the best barcode (prefer non-None barcode)
            if item.barcode and not data.get('barcode'):
                data['barcode'] = item.barcode

        # Filter prices for bodeguero role
        hide_prices = user.role == UserRole.BODEGUERO

        # Convert consolidated data to response items
        result = []

        # First add excluded items (they appear at the top for visibility)
        for item in excluded_items:
            result.append(InvoiceItemResponse(
                id=item.id,
                codigo_original=item.codigo_original,
                product_name=item.product_name,
                quantity=item.quantity,
                cantidad_original=item.cantidad_original,
                barcode=item.barcode,
                modified_by_bodeguero=item.modified_by_bodeguero,
                unit_price=None if hide_prices else item.unit_price,
                total_price=None if hide_prices else item.total_price,
                manual_sale_price=None if hide_prices else item.manual_sale_price,
                is_excluded=True,
                excluded_reason=item.excluded_reason,
                sync_success=item.sync_success,
                sync_error=item.sync_error,
                product_id=item.product_id,
                source_item_ids=None
            ))

        # Then add consolidated active items
        for key, data in consolidated.items():
            first_item = data['first_item']
            total_qty = data['total_quantity']
            total_price = data['total_price']

            # Calculate unit price from total/quantity (source of truth)
            unit_price = total_price / total_qty if total_qty > 0 and total_price > 0 else first_item.unit_price

            # Determine sync status for consolidated item
            if data['all_synced'] and len(data['items']) > 0 and all(i.sync_success for i in data['items']):
                sync_success = True
            elif data['any_failed']:
                sync_success = False
            else:
                sync_success = None

            # Get manual sale price (use first one if all are the same)
            manual_prices = data['manual_sale_prices']
            if manual_prices and all(p == manual_prices[0] for p in manual_prices):
                manual_sale_price = manual_prices[0]
            else:
                manual_sale_price = None

            result.append(InvoiceItemResponse(
                id=first_item.id,  # Use first item's ID as reference
                codigo_original=first_item.codigo_original,
                product_name=first_item.product_name,
                quantity=total_qty,
                cantidad_original=data['total_cantidad_original'],
                barcode=data.get('barcode') or first_item.barcode,
                modified_by_bodeguero=data['any_modified'],
                unit_price=None if hide_prices else unit_price,
                total_price=None if hide_prices else (total_price if total_price > 0 else None),
                manual_sale_price=None if hide_prices else manual_sale_price,
                is_excluded=False,
                excluded_reason=None,
                sync_success=sync_success,
                sync_error="; ".join(data['sync_errors']) if data['sync_errors'] else None,
                product_id=list(data['product_ids'])[0] if data['product_ids'] else None,
                source_item_ids=data['source_item_ids'] if len(data['source_item_ids']) > 1 else None
            ))

        return result

    def _item_to_response(
        self,
        item: PendingInvoiceItem,
        user: UserInfo
    ) -> InvoiceItemResponse:
        """Convert item to response with price filtering."""
        # Filter prices for bodeguero
        unit_price = item.unit_price if user.role != UserRole.BODEGUERO else None
        total_price = item.total_price if user.role != UserRole.BODEGUERO else None
        manual_sale_price = item.manual_sale_price if user.role != UserRole.BODEGUERO else None

        return InvoiceItemResponse(
            id=item.id,
            codigo_original=item.codigo_original,
            product_name=item.product_name,
            quantity=item.quantity,
            cantidad_original=item.cantidad_original,
            barcode=item.barcode,
            modified_by_bodeguero=item.modified_by_bodeguero,
            unit_price=unit_price,
            total_price=total_price,
            manual_sale_price=manual_sale_price,
            is_excluded=item.is_excluded,
            excluded_reason=item.excluded_reason,
            sync_success=item.sync_success,
            sync_error=item.sync_error,
            product_id=item.product_id
        )

    def _history_to_response(self, history: InvoiceHistory) -> InvoiceHistoryResponse:
        """Convert history to response."""
        items = [
            InvoiceHistoryItemResponse(
                id=item.id,
                codigo_original=item.codigo_original,
                barcode=item.barcode,
                product_id=item.product_id,
                product_name=item.product_name,
                quantity=item.quantity,
                unit_price=item.unit_price,
                total_value=item.total_value,
                success=item.success,
                error_message=item.error_message,
                was_modified=item.was_modified
            )
            for item in history.items
        ]

        return InvoiceHistoryResponse(
            id=history.id,
            invoice_number=history.invoice_number,
            supplier_name=history.supplier_name,
            uploaded_by=history.uploaded_by,
            synced_by=history.synced_by,
            synced_at=history.synced_at,
            total_items=history.total_items,
            successful_items=history.successful_items,
            failed_items=history.failed_items,
            has_errors=history.has_errors,
            items=items
        )

    # ========================================================================
    # LEGACY METHODS - Excel workflow (deprecated)
    # ========================================================================

    @staticmethod
    def extract_productos_from_xmls(xml_files: List[Dict[str, str]]) -> tuple[List[Dict[str, Any]], str]:
        """
        Extract unique products from multiple XML files.
        Sums quantities for products with the same codigo.

        DEPRECATED: Use upload_and_create_pending_invoices instead

        Args:
            xml_files: List of dicts with 'filename' and 'content'

        Returns:
            Tuple of (productos list, unified_xml string)
        """
        productos_map = {}  # Use dict to track unique products by codigo

        for xml_data in xml_files:
            content = xml_data['content']
            productos = extract_productos_from_xml(content)

            for producto in productos:
                codigo = producto['codigo']
                cantidad = producto.get('cantidad', 0)

                if codigo not in productos_map:
                    # First occurrence: store with cantidad
                    productos_map[codigo] = {
                        'codigo': producto['codigo'],
                        'descripcion': producto['descripcion'],
                        'cantidad': cantidad
                    }
                else:
                    # Duplicate: sum quantities
                    productos_map[codigo]['cantidad'] += cantidad

        # Convert map to list
        all_productos = list(productos_map.values())

        # Create unified XML
        unified_xml = create_unified_xml(xml_files)

        return all_productos, unified_xml

    @staticmethod
    def generate_excel(productos: List[Dict[str, Any]]) -> bytes:
        """
        Generate Excel file from productos list.

        DEPRECATED: Use new pending invoices workflow instead

        Args:
            productos: List of product dictionaries

        Returns:
            Excel file as bytes
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Productos"

        # Headers
        headers = ['CÓDIGO', 'DESCRIPCIÓN', 'CANTIDAD', 'CÓDIGO DE BARRAS']
        sheet.append(headers)

        # Style headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Add productos
        for producto in productos:
            sheet.append([
                producto['codigo'],
                producto['descripcion'],
                producto['cantidad'],
                ''  # Empty barcode column
            ])

        # Set column widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 70
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 20

        # Format cantidad column (integer format)
        for row in range(2, len(productos) + 2):
            cell = sheet.cell(row=row, column=3)
            cell.number_format = '0'

        # Format barcode column as text
        for row in range(2, len(productos) + 2):
            cell = sheet.cell(row=row, column=4)
            cell.number_format = '@'

        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    @staticmethod
    def update_xmls_with_barcodes(unified_xml: str, excel_data: List[List[Any]]) -> List[Dict[str, str]]:
        """
        Update XMLs with barcodes from Excel data.

        DEPRECATED: Use new pending invoices workflow instead

        Args:
            unified_xml: Unified XML content
            excel_data: Excel rows as list of lists

        Returns:
            List of updated XML files with 'filename' and 'content'
        """
        # Build codigo_map from Excel data
        # Row 0 is headers, so start from row 1
        codigo_map = {}

        for row in excel_data[1:]:  # Skip header row
            if len(row) < 4:
                continue

            codigo = str(row[0]).strip() if row[0] else None
            cantidad = float(row[2]) if row[2] else 0  # Column C: cantidad
            codigo_barras = str(row[3]).strip() if row[3] else None

            if codigo:
                # If barcode is empty, use original codigo (no barcode change, only cantidad update)
                if not codigo_barras or codigo_barras == '':
                    codigo_barras = codigo

                # Store barcode and cantidad
                data = {
                    'barcode': codigo_barras,
                    'cantidad': cantidad
                }
                codigo_map[codigo] = data
                codigo_map[codigo + ' '] = data  # Handle trailing space in XML

        # Update XMLs with consolidated option (single XML with all products)
        updated_xmls = update_xml_with_barcodes_consolidated(unified_xml, codigo_map)

        return updated_xmls
